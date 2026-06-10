from __future__ import annotations

import ast
import json
import operator
import os
import re
import sys
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


SUMMARY_WORDS = {"小计", "安装服务费", "运维服务费", "含税总计", "总计", "设备含税小计"}
CODE_RE = re.compile(r"^(?:[A-Z]{2}_[A-Z0-9]{2,4}\d{3,5}|[A-Z]{2}\d{4}|RJ\d{4})$")


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[ \t]+", " ", str(value).replace("\r\n", "\n").replace("\r", "\n")).strip()


def is_code(value: Any) -> bool:
    return bool(CODE_RE.fullmatch(clean(value)))


def round_half_up(value: float, digits: int = 0) -> float:
    quantum = Decimal("1") if digits == 0 else Decimal("1").scaleb(-digits)
    return float(Decimal(str(value)).quantize(quantum, rounding=ROUND_HALF_UP))


class WorkbookEvaluator:
    def __init__(self, workbook: Any):
        self.workbook = workbook
        self.cache: dict[tuple[str, str], Any] = {}

    def cell(self, sheet_name: str, ref: str) -> Any:
        key = (sheet_name, ref.upper())
        if key in self.cache:
            return self.cache[key]
        worksheet = self.workbook[sheet_name]
        value = worksheet[ref].value
        self.cache[key] = None
        result = self.value(sheet_name, value)
        self.cache[key] = result
        return result

    def value(self, sheet_name: str, value: Any) -> Any:
        if isinstance(value, (int, float)):
            return value
        if value is None:
            return None
        text = clean(value)
        if not text:
            return None
        if not text.startswith("="):
            try:
                return float(text)
            except ValueError:
                return text
        return self.formula(sheet_name, text[1:])

    def formula(self, current_sheet: str, expr: str) -> Any:
        expr = expr.strip()
        expr = self._replace_functions(current_sheet, expr)
        expr = self._replace_refs(current_sheet, expr)
        return safe_eval(expr)

    def _range_values(self, current_sheet: str, token: str) -> list[Any]:
        sheet_name, range_ref = split_sheet_ref(current_sheet, token)
        start_ref, end_ref = range_ref.split(":", 1)
        start_col, start_row = split_cell_ref(start_ref)
        end_col, end_row = split_cell_ref(end_ref)
        values: list[Any] = []
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ref = f"{column_letter(col)}{row}"
                values.append(self.cell(sheet_name, ref))
        return values

    def _cell_value(self, current_sheet: str, token: str) -> Any:
        sheet_name, cell_ref = split_sheet_ref(current_sheet, token)
        return self.cell(sheet_name, cell_ref)

    def _replace_functions(self, current_sheet: str, expr: str) -> str:
        while True:
            match = re.search(r"\b(SUM|COUNTA|ROUND)\(", expr, re.IGNORECASE)
            if not match:
                return expr
            name = match.group(1).upper()
            open_index = match.end() - 1
            close_index = find_matching_paren(expr, open_index)
            args = split_args(expr[open_index + 1 : close_index])
            if name == "SUM":
                total = 0.0
                for arg in args:
                    if ":" in arg:
                        total += sum(number(v) for v in self._range_values(current_sheet, arg))
                    else:
                        total += number(self._eval_arg(current_sheet, arg))
                replacement = str(total)
            elif name == "COUNTA":
                count = 0
                for arg in args:
                    values = self._range_values(current_sheet, arg) if ":" in arg else [self._eval_arg(current_sheet, arg)]
                    count += sum(1 for value in values if clean(value) != "")
                replacement = str(count)
            else:
                value = number(self._eval_arg(current_sheet, args[0]))
                digits = int(number(self._eval_arg(current_sheet, args[1]))) if len(args) > 1 else 0
                replacement = str(round_half_up(value, digits))
            expr = expr[: match.start()] + replacement + expr[close_index + 1 :]

    def _replace_refs(self, current_sheet: str, expr: str) -> str:
        ref_pattern = re.compile(
            r"(?:(?:'([^']+)'|([^+\-*/(),\s!]+))!)?(\$?[A-Z]{1,3}\$?\d+)",
            re.IGNORECASE,
        )

        def replace(match: re.Match[str]) -> str:
            full = match.group(0)
            start = match.start()
            if start > 0 and (expr[start - 1].isalnum() or expr[start - 1] == "_"):
                return full
            sheet_name = match.group(1) or match.group(2) or current_sheet
            ref = match.group(3).replace("$", "")
            return str(number(self.cell(sheet_name, ref)))

        return ref_pattern.sub(replace, expr)

    def _eval_arg(self, current_sheet: str, arg: str) -> Any:
        arg = arg.strip()
        if re.fullmatch(r"(?:'[^']+'|[^+\-*/(),\s!]+)!\$?[A-Z]{1,3}\$?\d+|\$?[A-Z]{1,3}\$?\d+", arg, re.I):
            return self._cell_value(current_sheet, arg.replace("$", ""))
        return self.formula(current_sheet, arg)


def split_args(text: str) -> list[str]:
    args: list[str] = []
    depth = 0
    start = 0
    for index, char in enumerate(text):
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        elif char == "," and depth == 0:
            args.append(text[start:index].strip())
            start = index + 1
    args.append(text[start:].strip())
    return args


def find_matching_paren(text: str, open_index: int) -> int:
    depth = 0
    for index in range(open_index, len(text)):
        if text[index] == "(":
            depth += 1
        elif text[index] == ")":
            depth -= 1
            if depth == 0:
                return index
    raise ValueError(f"Unmatched parenthesis in formula: {text}")


def split_sheet_ref(current_sheet: str, token: str) -> tuple[str, str]:
    token = token.replace("$", "").strip()
    if "!" not in token:
        return current_sheet, token
    sheet, ref = token.rsplit("!", 1)
    return sheet.strip("'"), ref


def split_cell_ref(ref: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]{1,3})(\d+)", ref.replace("$", "").upper())
    if not match:
        raise ValueError(f"Invalid cell ref: {ref}")
    return column_index_from_string(match.group(1)), int(match.group(2))


def column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(clean(value))
    except ValueError:
        return 0.0


def safe_eval(expr: str) -> float:
    allowed = {
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Constant,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.USub,
        ast.UAdd,
    }
    ops = {
        ast.Add: operator.add,
        ast.Sub: operator.sub,
        ast.Mult: operator.mul,
        ast.Div: operator.truediv,
        ast.USub: operator.neg,
        ast.UAdd: operator.pos,
    }

    def walk(node: ast.AST) -> float:
        if type(node) not in allowed:
            raise ValueError(f"Unsupported expression: {expr}")
        if isinstance(node, ast.Expression):
            return walk(node.body)
        if isinstance(node, ast.Constant):
            return float(node.value)
        if isinstance(node, ast.BinOp):
            return ops[type(node.op)](walk(node.left), walk(node.right))
        if isinstance(node, ast.UnaryOp):
            return ops[type(node.op)](walk(node.operand))
        raise ValueError(f"Unsupported expression: {expr}")

    return walk(ast.parse(expr, mode="eval"))


def find_sheet(workbook: Any) -> Any:
    for worksheet in workbook.worksheets:
        headers = [clean(worksheet.cell(2, column).value) for column in range(1, worksheet.max_column + 1)]
        if "设备名称" in headers and "料号" in headers:
            return worksheet
    raise RuntimeError("Unable to find detail sheet with 设备名称 and 料号")


def export_items(workbook_path: Path, out_json: Path, report_json: Path) -> None:
    workbook = load_workbook(workbook_path, data_only=False)
    evaluator = WorkbookEvaluator(workbook)
    worksheet = find_sheet(workbook)

    rows: list[dict[str, Any]] = []
    groups: dict[str, dict[str, Any]] = {}
    price_conflicts: list[dict[str, Any]] = []
    row_warnings: list[dict[str, Any]] = []
    aggregate_duplicates = os.environ.get("A8_AGGREGATE_DUPLICATE_MATERIALS", "").lower() in {"1", "true", "yes", "y"}

    for row in range(3, worksheet.max_row + 1):
        code = clean(worksheet.cell(row, 14).value)
        device = clean(worksheet.cell(row, 3).value)
        first = clean(worksheet.cell(row, 1).value)
        qty = evaluator.value(worksheet.title, worksheet.cell(row, 7).value)
        tax_price = evaluator.value(worksheet.title, worksheet.cell(row, 10).value)
        if first in SUMMARY_WORDS or device in SUMMARY_WORDS:
            continue
        if not (code or clean(qty) or clean(tax_price)):
            continue
        qty_num = number(qty)
        price_num = number(tax_price)
        blank_fields = []
        if not code:
            blank_fields.append("material code")
        elif not is_code(code):
            row_warnings.append({"type": "material-code-pattern", "row": row, "code": code})
        if clean(qty) == "":
            blank_fields.append("quantity")
        if clean(tax_price) == "":
            blank_fields.append("tax price")
        if blank_fields:
            row_warnings.append({"type": "blank-fields-preserved", "row": row, "code": code, "blankFields": blank_fields})

        row_item = {
            "sourceRow": row,
            "code": code,
            "label": device or code or f"row-{row}",
            "brand": clean(worksheet.cell(row, 4).value),
            "model": clean(worksheet.cell(row, 5).value),
            "qty": "" if clean(qty) == "" else qty_num,
            "taxPrice": "" if clean(tax_price) == "" else price_num,
        }
        rows.append(row_item)
        if not is_code(code):
            continue

        group = groups.setdefault(
            code,
            {
                "code": code,
                "qty": 0.0,
                "taxPrice": price_num,
                "sourceRows": [],
                "deviceNames": [],
            },
        )
        if abs(group["taxPrice"] - price_num) > 0.0001:
            price_conflicts.append(
                {
                    "code": code,
                    "existingTaxPrice": group["taxPrice"],
                    "newTaxPrice": price_num,
                    "row": row,
                    "device": device,
                }
            )
        group["qty"] += qty_num
        group["sourceRows"].append(row)
        if device not in group["deviceNames"]:
            group["deviceNames"].append(device)

    duplicate_codes = [
        {"code": code, "sourceRows": group["sourceRows"]}
        for code, group in sorted(groups.items())
        if len(group["sourceRows"]) > 1
    ]
    for duplicate in duplicate_codes:
        row_warnings.append(
            {
                "type": "duplicate-material-code-preserved",
                "code": duplicate["code"],
                "sourceRows": duplicate["sourceRows"],
            }
        )

    if aggregate_duplicates:
        items = []
        for code in sorted(groups):
            group = groups[code]
            qty = group["qty"]
            items.append(
                {
                    "code": code,
                    "qty": int(qty) if abs(qty - round(qty)) < 0.0001 else round(qty, 4),
                    "taxPrice": round(group["taxPrice"], 4),
                    "sourceRows": group["sourceRows"],
                    "label": " / ".join(group["deviceNames"]),
                }
            )
    else:
        items = rows

    payload = {
        "workbook": str(workbook_path),
        "sheet": worksheet.title,
        "groupBy": "code" if aggregate_duplicates else "none",
        "aggregateDuplicates": aggregate_duplicates,
        "items": items,
    }
    report = {
        "workbook": str(workbook_path),
        "sheet": worksheet.title,
        "rawMatchedRows": len(rows),
        "exportedItems": len(items),
        "aggregatedItems": len(items) if aggregate_duplicates else None,
        "aggregateDuplicates": aggregate_duplicates,
        "duplicateCodes": duplicate_codes,
        "priceConflicts": price_conflicts,
        "warnings": row_warnings,
        "rows": rows,
        "items": items,
    }
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(payload["items"], ensure_ascii=False, indent=2), encoding="utf-8")
    report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    workbook.close()
    print(
        json.dumps(
            {
                "outJson": str(out_json),
                "reportJson": str(report_json),
                "rawMatchedRows": len(rows),
                "exportedItems": len(items),
                "aggregateDuplicates": aggregate_duplicates,
                "duplicateCodes": len(duplicate_codes),
                "warnings": len(row_warnings),
                "priceConflicts": len(price_conflicts),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def main() -> int:
    if len(sys.argv) != 4:
        print("Usage: export-init-material-items.py <filled.xlsx> <items.json> <report.json>", file=sys.stderr)
        return 2
    export_items(Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
